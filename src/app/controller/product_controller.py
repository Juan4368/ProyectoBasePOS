from io import BytesIO
from typing import Annotated

from openpyxl import load_workbook
from pydantic import ValidationError

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from src.config import get_db
from src.domain.dtos.genericResponseDto import CreationResponse
from src.domain.dtos.productsDto import (
    ProductRequest,
    ProductResponse,
    ProductImportError,
    ProductImportResponse,
    ProductStatusUpdate,
)
from src.domain.services.product_service import ProductService
from src.infrastructure.repository.createProductsRepository import ProductRepository

router = APIRouter(prefix="/productos", tags=["productos"])


def get_product_service(db: Session = Depends(get_db)) -> ProductService:
    repo = ProductRepository(db)
    return ProductService(repo)


DbDep = Annotated[Session, Depends(get_db)]
ServiceDep = Annotated[ProductService, Depends(get_product_service)]


@router.post(
    "/",
    response_model=CreationResponse[ProductResponse],
    status_code=status.HTTP_201_CREATED,
)
def create_product(
    payload: ProductRequest,
    service: ServiceDep,
) -> CreationResponse[ProductResponse]:
    try:
        created = service.create_product(payload)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)
        ) from exc

    return CreationResponse[ProductResponse](id=created.producto_id, data=created)


@router.get("/", response_model=list[ProductResponse])
def list_products(service: ServiceDep) -> list[ProductResponse]:
    return service.list_products()


@router.get("/buscar", response_model=list[ProductResponse])
def search_products(q: str, service: ServiceDep) -> list[ProductResponse]:
    if not q.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El parametro q no puede estar vacio",
        )
    return service.search_products(q.strip())


@router.get("/{product_id}", response_model=ProductResponse)
def get_product(product_id: int, service: ServiceDep) -> ProductResponse:
    producto = service.get_product(product_id)
    if not producto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Producto no encontrado",
        )
    return producto


@router.post("/import", response_model=ProductImportResponse)
async def import_products(
    service: ServiceDep,
    file: UploadFile = File(...),
) -> ProductImportResponse:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser un .xlsx",
        )

    try:
        content = await file.read()
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se pudo leer el archivo: {exc}",
        ) from exc

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo no tiene encabezados",
        )

    field_names = list(ProductRequest.model_fields.keys())
    field_map = {name.lower(): name for name in field_names}
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    header_map = {idx: field_map.get(h.lower()) for idx, h in enumerate(headers)}

    required_fields = [
        name for name, field in ProductRequest.model_fields.items() if field.is_required()
    ]
    missing = [name for name in required_fields if name not in header_map.values()]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas requeridas: {', '.join(missing)}",
        )

    items: list[ProductRequest] = []
    errors: list[ProductImportError] = []
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        payload = {}
        for idx, value in enumerate(row):
            field_name = header_map.get(idx)
            if not field_name:
                continue
            if isinstance(value, str):
                value = value.strip()
                if value == "":
                    value = None
            payload[field_name] = value

        try:
            items.append(ProductRequest(**payload))
        except ValidationError as exc:
            errors.append(
                ProductImportError(row=row_index, message=str(exc.errors()))
            )

    created, skipped = service.import_products(items)
    return ProductImportResponse(
        created=created,
        skipped=skipped,
        invalid=len(errors),
        errors=errors,
    )


@router.put("/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: int,
    payload: ProductRequest,
    service: ServiceDep,
) -> ProductResponse:
    try:
        updated = service.update_product(product_id, payload)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)
        ) from exc

    if not updated:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Producto no encontrado",
        )
    return updated


@router.patch("/{product_id}/estado", response_model=ProductResponse)
def update_product_status(
    product_id: int,
    payload: ProductStatusUpdate,
    service: ServiceDep,
) -> ProductResponse:
    try:
        updated = service.update_product_status(product_id, payload)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)
        ) from exc

    if not updated:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Producto no encontrado",
        )
    return updated
